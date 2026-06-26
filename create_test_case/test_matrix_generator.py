"""
测试矩阵生成模块

负责读取test case JSON，本地生成Excel格式的测试矩阵
支持独立运行和被main.py调用两种方式
"""

import os
import sys
import re
import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 解决 Windows 控制台中文打印乱码问题
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# 颜色常量
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E75", fill_type="solid")  # 深蓝表头
HEADER_FONT = Font(color="FFFFFF", bold=True)
HAPPY_PATH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色
BOUNDARY_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
TIMEOUT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # 蓝色
INTERRUPT_FILL = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")  # 紫色
SUMMARY_ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")  # 浅蓝交替行

# 测试区域关键词匹配规则
TEST_AREA_KEYWORDS = {
    "Negative": ["negative", "wrong", "invalid", "error", "[tbd]", "incorrect", "unexpected"],
    "Boundary": ["boundary", "counter", "limit", "count", "edge", "exactly", "minimum", "maximum"],
    "Timeout": ["timeout", "idle", "sleep", "wake", "delay", "sec"],
    "Interrupt": ["interrupt", "reset", "battery", "exception", "crash", "power"],
}

# 各测试区域对应的列字母
AREA_COLUMNS = {
    "Happy Path": "B",  # Test Description
    "Negative": "B",
    "Boundary": "B",
    "Timeout": "B",
    "Interrupt": "B",
}


def _classify_test_area(test_case):
    """
    根据测试用例内容判断测试区域

    :param test_case: 测试用例字典
    :return: 测试区域名称
    """
    description = (test_case.get("Test Description", "") or "").lower()
    expected = (test_case.get("Expected Result", "") or "").lower()
    combined = description + " " + expected

    # 检查是否有[TBD]标记（Negative区域的特征）
    if "[tbd]" in combined:
        return "Negative"

    # 按优先级匹配关键词
    for area, keywords in TEST_AREA_KEYWORDS.items():
        for keyword in keywords:
            if keyword in combined:
                return area

    # 默认为Happy Path
    return "Happy Path"


def _get_area_fill(area_name):
    """
    根据测试区域名称获取对应的填充颜色

    :param area_name: 测试区域名称
    :return: PatternFill对象
    """
    fill_map = {
        "Happy Path": HAPPY_PATH_FILL,
        "Negative": NEGATIVE_FILL,
        "Boundary": BOUNDARY_FILL,
        "Timeout": TIMEOUT_FILL,
        "Interrupt": INTERRUPT_FILL,
    }
    return fill_map.get(area_name, None)


class TestMatrixGenerator:
    """测试矩阵生成器"""

    def __init__(self):
        """初始化测试矩阵生成器"""
        pass

    def _get_next_version(self, output_dir):
        """
        获取下一个版本号

        :param output_dir: 输出目录
        :return: 下一个版本号（整数）
        """
        version_pattern = re.compile(r"test_matrix_v(\d+)\.xlsx")
        existing_versions = []

        if os.path.exists(output_dir):
            for filename in os.listdir(output_dir):
                match = version_pattern.match(filename)
                if match:
                    existing_versions.append(int(match.group(1)))

        if not existing_versions:
            return 1
        else:
            return max(existing_versions) + 1

    def _sanitize_sheet_name(self, name):
        """
        清理sheet名称（Excel限制）
        Excel sheet名称不能包含: []:*?/\

        :param name: 原始名称
        :return: 清理后的名称（最多31个字符）
        """
        safe_name = str(name)[:31]
        for char in "[]:*?/\\":
            safe_name = safe_name.replace(char, "")
        return safe_name

    def _apply_cell_style(self, cell, fill=None, font=None, alignment=None):
        """
        应用单元格样式

        :param cell: 单元格对象
        :param fill: 填充颜色
        :param font: 字体
        :param alignment: 对齐方式
        """
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment

    def _create_summary_sheet(self, wb, modes_data):
        """
        创建Summary sheet

        :param wb: Workbook对象
        :param modes_data: modes_data字典
        """
        ws = wb.active
        ws.title = "Summary"

        # 标题行
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "QA Test Matrix Summary"
        title_cell.fill = HEADER_FILL
        title_cell.font = HEADER_FONT
        title_cell.alignment = Alignment(horizontal="center")

        # 表头
        headers = ["Mode", "Total", "Executed", "Passed", "Failed", "Skipped"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        row = 3
        for mode_name in modes_data.keys():
            # 清理sheet名称（与数据表保持一致）
            safe_name = self._sanitize_sheet_name(mode_name)

            # Mode名称（显示清理后的名称）
            ws.cell(row=row, column=1, value=safe_name)

            # Total - 使用COUNT公式
            total_formula = f"=COUNT('{safe_name}'!A2:A100)"
            ws.cell(row=row, column=2, value=total_formula)

            # Executed - 使用COUNTIF公式（Passed + Failed + Skipped）
            executed_formula = f'=COUNTIF(\'{safe_name}\'!D2:D100,"Passed")+COUNTIF(\'{safe_name}\'!D2:D100,"Failed")+COUNTIF(\'{safe_name}\'!D2:D100,"Skipped")'
            ws.cell(row=row, column=3, value=executed_formula)

            # Passed
            passed_formula = f'=COUNTIF(\'{safe_name}\'!D2:D100,"Passed")'
            ws.cell(row=row, column=4, value=passed_formula)

            # Failed
            failed_formula = f'=COUNTIF(\'{safe_name}\'!D2:D100,"Failed")'
            ws.cell(row=row, column=5, value=failed_formula)

            # Skipped
            skipped_formula = f'=COUNTIF(\'{safe_name}\'!D2:D100,"Skipped")'
            ws.cell(row=row, column=6, value=skipped_formula)

            # 交替行颜色
            if row % 2 == 1:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = SUMMARY_ALT_FILL

            row += 1

        # TOTAL行
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        for col in range(2, 7):
            col_letter = chr(64 + col)  # A=65, B=66, ...
            ws.cell(row=row, column=col, value=f"=SUM({col_letter}3:{col_letter}{row-1})")
            ws.cell(row=row, column=col).font = Font(bold=True)

        # 设置列宽
        ws.column_dimensions["A"].width = 20
        for col in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 12

    def _create_mode_sheet(self, wb, mode_name, test_cases):
        """
        创建Mode sheet

        :param wb: Workbook对象
        :param mode_name: Mode名称
        :param test_cases: 测试用例列表
        """
        # 清理sheet名称（Excel限制）
        safe_name = self._sanitize_sheet_name(mode_name)
        ws = wb.create_sheet(title=safe_name)

        # 表头
        headers = [
            "Test ID", "Test Description", "Expected Result",
            "Result (Passed/Failed/Skipped)", "Comment",
            "Internal Bug ID", "External Bug ID"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        row = 2
        for test_case in test_cases:
            # 判断测试区域
            area = _classify_test_area(test_case)
            area_fill = _get_area_fill(area)

            # 写入数据
            values = [
                test_case.get("Test ID", ""),
                test_case.get("Test Description", ""),
                test_case.get("Expected Result", ""),
                test_case.get("Results", ""),
                test_case.get("Comment", ""),
                test_case.get("Internal Bug ID", ""),
                test_case.get("External Bug ID", ""),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                # 对Test Description列应用颜色
                if col == 2 and area_fill:
                    cell.fill = area_fill

            row += 1

        # 设置列宽
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15

    def generate_excel(self, test_cases_json_path, output_dir):
        """
        读取test case JSON，本地生成Excel格式的测试矩阵

        :param test_cases_json_path: test case JSON文件路径
        :param output_dir: 输出目录
        :return: 生成的文件路径，失败返回None
        """
        # 检查JSON文件是否存在
        if not os.path.exists(test_cases_json_path):
            print(f"错误: test case JSON文件不存在: {test_cases_json_path}")
            return None

        # 读取JSON文件
        with open(test_cases_json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # 支持两种格式：对象格式(modes_data)和数组格式
        if isinstance(data, dict) and "modes_data" in data:
            modes_data = data["modes_data"]
        elif isinstance(data, list):
            # 数组格式：所有test case放到一个Default sheet
            modes_data = {"Default": data}
        else:
            print("错误: JSON格式不支持，需要modes_data结构或数组")
            return None

        if not modes_data:
            print("错误: 没有测试用例数据")
            return None

        # 创建Excel工作簿
        wb = Workbook()

        # 创建Summary sheet
        self._create_summary_sheet(wb, modes_data)

        # 创建各Mode sheet
        for mode_name, test_cases in modes_data.items():
            if test_cases:
                self._create_mode_sheet(wb, mode_name, test_cases)

        # 获取下一个版本号
        version = self._get_next_version(output_dir)
        output_filename = f"test_matrix_v{version}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        # 保存Excel文件
        wb.save(output_path)
        print(f"成功生成Excel测试矩阵: {output_filename}")
        return output_path


def main():
    """独立运行入口"""
    parser = argparse.ArgumentParser(description="测试矩阵生成工具 - 从test case JSON生成Excel测试矩阵")
    parser.add_argument("--input-json", required=True, help="test case JSON文件路径")
    parser.add_argument("--output-dir", required=True, help="输出目录路径")

    args = parser.parse_args()

    print("=" * 60)
    print("测试矩阵生成工具")
    print("=" * 60)
    print(f"输入JSON文件: {args.input_json}")
    print(f"输出目录: {args.output_dir}")
    print("=" * 60)

    # 创建测试矩阵生成器并运行
    generator = TestMatrixGenerator()
    result_path = generator.generate_excel(args.input_json, args.output_dir)

    if result_path:
        print("\n" + "=" * 60)
        print("测试矩阵生成完成！")
        print(f"生成的文件: {result_path}")
        print("=" * 60)
        return 0
    else:
        print("\n错误: 测试矩阵生成失败")
        return 1


if __name__ == "__main__":
    sys.exit(main())
